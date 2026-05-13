import pandas as pd
import instaloader
import smtplib
import socks
import socket
import random
import time
import os
from email.message import EmailMessage
from datetime import date
from stem import Signal
from stem.control import Controller

# --- CREDENTIALS & CONFIG ---
GMAIL_USER  = os.environ.get("GMAIL_USER")
GMAIL_PASS  = os.environ.get("GMAIL_PASS")
INSTA_USER  = os.environ.get("INSTA_USER")   # Optional but recommended
INSTA_PASS  = os.environ.get("INSTA_PASS")   # Optional but recommended
TOR_PASSWORD = os.environ.get("TOR_PASSWORD", "")  # Set if you configured a control password

RECEIVER_EMAIL = ["mhokijanto@gmail.com"]
INPUT_FILE     = "INSTA INFLUENCER VISIBILITY TABLE.xlsx"
OUTPUT_FILE    = f"Cimory_Reels_Report_{date.today()}.xlsx"

REELS_LIMIT   = 20
TOR_SOCKS_PORT = 9050   # Default Tor SOCKS port
TOR_CTRL_PORT  = 9051   # Default Tor control port

# Randomized delay range (seconds) between post fetches — mimics human browsing
MIN_DELAY = 3.0
MAX_DELAY = 7.0


# ---------------------------------------------------------------------------
# TOR HELPERS
# ---------------------------------------------------------------------------

def enable_tor_proxy():
    """Route all socket traffic through Tor's SOCKS5 proxy."""
    socks.set_default_proxy(socks.SOCKS5, "127.0.0.1", TOR_SOCKS_PORT)
    socket.socket = socks.socksocket
    print("🧅 Tor proxy enabled.")


def rotate_tor_ip():
    """
    Signal Tor to build a new circuit (= new exit IP).
    Requires the Tor control port to be open and optionally a password.
    """
    try:
        with Controller.from_port(port=TOR_CTRL_PORT) as ctrl:
            ctrl.authenticate(password=TOR_PASSWORD)
            ctrl.signal(Signal.NEWNYM)
        # Tor needs a moment to establish the new circuit
        time.sleep(3)
        print("🔄 Tor IP rotated.")
    except Exception as e:
        print(f"⚠️  Could not rotate Tor IP: {e} — continuing with current IP.")


def get_current_ip() -> str:
    """Check current exit IP via Tor (useful for debugging)."""
    try:
        import urllib.request
        proxy = urllib.request.ProxyHandler({
            "http":  f"socks5h://127.0.0.1:{TOR_SOCKS_PORT}",
            "https": f"socks5h://127.0.0.1:{TOR_SOCKS_PORT}",
        })
        opener = urllib.request.build_opener(proxy)
        ip = opener.open("https://api.ipify.org", timeout=10).read().decode()
        return ip
    except Exception:
        return "unknown"


# ---------------------------------------------------------------------------
# INSTALOADER
# ---------------------------------------------------------------------------

def build_loader() -> instaloader.Instaloader:
    """Create an Instaloader instance that routes through Tor."""
    L = instaloader.Instaloader(
        download_pictures=False,
        download_videos=False,
        download_video_thumbnails=False,
        download_geotags=False,
        download_comments=False,
        save_metadata=False,
        compress_json=False,
        quiet=True,
        # Randomize user-agent slightly to avoid fingerprinting
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
    )
    if INSTA_USER and INSTA_PASS:
        try:
            L.login(INSTA_USER, INSTA_PASS)
            print(f"✅ Logged in as {INSTA_USER}")
        except Exception as e:
            print(f"⚠️  Login failed ({e}), continuing anonymously.")
    else:
        print("ℹ️  Scraping anonymously via Tor.")
    return L


def scrape_reels(L: instaloader.Instaloader, username: str, limit: int = REELS_LIMIT) -> list[dict]:
    """Fetch up to `limit` most-recent video posts (Reels) for a given username."""
    try:
        profile = instaloader.Profile.from_username(L.context, username)
    except instaloader.exceptions.ProfileNotExistsException:
        print(f"❌ Profile not found: {username}")
        return []
    except Exception as e:
        print(f"❌ Could not load profile {username}: {e}")
        return []

    reels = []
    try:
        for post in profile.get_posts():
            if not post.is_video:
                continue
            reels.append({
                "views":    post.video_view_count or 0,
                "likes":    post.likes             or 0,
                "comments": post.comments          or 0,
                "url":      f"https://www.instagram.com/p/{post.shortcode}/",
            })
            if len(reels) >= limit:
                break
            # Randomized human-like delay between post fetches
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
    except Exception as e:
        print(f"⚠️  Stopped early for {username}: {e}")

    return reels


# ---------------------------------------------------------------------------
# METRICS
# ---------------------------------------------------------------------------

def compute_metrics(handle: str, reels: list[dict]) -> dict | None:
    if not reels:
        print(f"⚠️  No reels found for {handle}, skipping.")
        return None

    df = pd.DataFrame(reels).sort_values("views", ascending=False).reset_index(drop=True)

    avg_views    = df["views"].mean()
    avg_likes    = df["likes"].mean()
    avg_comments = df["comments"].mean()

    top5    = df.head(5)
    bottom5 = df.tail(5)

    er = round((avg_likes + avg_comments) / avg_views * 100, 2) if avg_views > 0 else 0.0

    return {
        "Instagram Handle":    handle,
        "Reels Analyzed":      len(df),
        "Avg Views":           round(avg_views),
        "Avg Likes":           round(avg_likes),
        "Avg Comments":        round(avg_comments),
        "Avg Top 5 Views":     round(top5["views"].mean()),
        "Avg Bottom 5 Views":  round(bottom5["views"].mean()),
        "Most Views":          int(df["views"].max()),
        "Engagement Rate (%)": er,
        "Top Reels Links":     "\n".join(top5["url"].tolist()),
        "Bottom Reels Links":  "\n".join(bottom5["url"].tolist()),
    }


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------

def export_excel(results: list[dict], output_path: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(results)
    df.to_excel(output_path, index=False, sheet_name="Reels Report")

    wb = load_workbook(output_path)
    ws = wb["Reels Report"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    link_cols = {"Top Reels Links", "Bottom Reels Links"}
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header_val = ws.cell(row=1, column=col_idx).value or ""
        if header_val in link_cols:
            ws.column_dimensions[get_column_letter(col_idx)].width = 55
        else:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")


# ---------------------------------------------------------------------------
# EMAIL
# ---------------------------------------------------------------------------

def send_email(output_path: str) -> None:
    if not GMAIL_USER or not GMAIL_PASS:
        print("⚠️  Email credentials not set — skipping email.")
        return

    msg = EmailMessage()
    msg["Subject"] = f"📊 CIMORY: Reels Visibility Study — {date.today()}"
    msg["From"]    = GMAIL_USER
    msg["To"]      = ", ".join(RECEIVER_EMAIL)
    msg.set_content(
        "Hi Clarissa,\n\n"
        "Please find attached the automated Reels Visibility Study.\n\n"
        "Generated by Cimory Bot 5.0"
    )
    with open(output_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(output_path),
        )
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_PASS)
            server.send_message(msg)
        print("📧 Email sent successfully!")
    except Exception as e:
        print(f"❌ Email failed: {e}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def run():
    print("🚀 Cimory Instagram Reels Tracker v5.0 starting...")

    # Load input
    try:
        df_input = pd.read_excel(INPUT_FILE)
        handles  = (
            df_input["Instagram Handle"]
            .dropna()
            .astype(str)
            .str.replace("@", "")
            .str.strip()
            .tolist()
        )
        print(f"📋 Loaded {len(handles)} influencers from {INPUT_FILE}")
    except Exception as e:
        print(f"❌ Failed to read input file: {e}")
        return

    # Enable Tor proxy for all subsequent traffic
    enable_tor_proxy()
    print(f"🌐 Starting IP: {get_current_ip()}")

    results = []
    for i, handle in enumerate(handles, start=1):
        # Rotate IP before each new profile
        print(f"\n[{i}/{len(handles)}] Rotating IP before scraping @{handle}...")
        rotate_tor_ip()
        print(f"   🌐 New IP: {get_current_ip()}")

        # Fresh loader per profile so sessions don't carry over
        L = build_loader()

        reels   = scrape_reels(L, handle)
        metrics = compute_metrics(handle, reels)
        if metrics:
            results.append(metrics)
            print(f"   ✔  {len(reels)} reels | Avg Views: {metrics['Avg Views']:,} | ER: {metrics['Engagement Rate (%)']:.2f}%")

        # Extra cooldown between profiles (30–60s) — looks more human
        cooldown = random.uniform(30, 60)
        print(f"   ⏳ Cooling down {cooldown:.0f}s before next profile...")
        time.sleep(cooldown)

    if not results:
        print("⚠️  No data collected. Exiting.")
        return

    export_excel(results, OUTPUT_FILE)
    send_email(OUTPUT_FILE)
    print("\n🏁 Done!")


if __name__ == "__main__":
    run()
